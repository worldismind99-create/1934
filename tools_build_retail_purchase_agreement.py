# -*- coding: utf-8 -*-
"""Blank RETAIL PURCHASE AGREEMENT form, sized to fill a single A4 page."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.worksheet.page import PageMargins

OUT = "/home/user/1934/Retail_Purchase_Agreement_Blank.xlsx"
EMBLEM = "/home/user/1934/assets/emblem_sample_ford.png"
EMU = 9525  # EMU per pixel

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Retail Purchase Agreement"

FONT = "Arial"
thin = Side(style="thin", color="000000")
medium = Side(style="medium", color="000000")
box = Border(left=thin, right=thin, top=thin, bottom=thin)
underline = Border(bottom=thin)
MONEY = '$#,##0.00;($#,##0.00);"-"'


def f(size, bold=False, italic=False):
    return Font(name=FONT, size=size, bold=bold, italic=italic)


def norm(rng):
    t = ws[rng]
    if not isinstance(t, tuple):
        return ((t,),)
    if not isinstance(t[0], tuple):
        return (t,)
    return t


def put(cell, value=None, size=8, bold=False, italic=False, align="left",
        valign="center", wrap=False, merge=None, border=None, fmt=None, shrink=False):
    if merge:
        ws.merge_cells(merge)
    c = ws[cell]
    if value is not None:
        c.value = value
    c.font = f(size, bold, italic)
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap,
                            shrink_to_fit=shrink)
    if fmt:
        c.number_format = fmt
    if border:
        for row in norm(merge if merge else cell):
            for cc in row:
                cc.border = border
    return c


def outline(rng, side=thin):
    cells = norm(rng)
    nrows, ncols = len(cells), len(cells[0])
    for i, row in enumerate(cells):
        for j, c in enumerate(row):
            b = c.border
            c.border = Border(
                left=side if j == 0 else b.left,
                right=side if j == ncols - 1 else b.right,
                top=side if i == 0 else b.top,
                bottom=side if i == nrows - 1 else b.bottom,
            )


# ------------------------------------------------------- columns (A4 width)
# total ≈ 751 px ≈ 198.7 mm; A4 (210 mm) less 2 × 0.22" margins ≈ 198.8 mm
widths = {"A": 1.2, "B": 7.1, "C": 7.1, "D": 7.1, "E": 7.1, "F": 7.1,
          "G": 7.1, "H": 7.1, "I": 7.1, "J": 9.0, "K": 20.0, "L": 9.5, "M": 1.2}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ------------------------------------------------------- dealer header
put("B1", "EXCELLENT AUTO SALES AND LEASING", size=10, bold=True, merge="B1:F1")
put("B2", "6419 ALONDRA BLVD", size=8.5, merge="B2:F2")
put("B3", "PARAMOUNT, CA 90723", size=8.5, merge="B3:F3")
put("B4", "Tel: (310)906-8117      Fax: (562)529-7997", size=8.5, merge="B4:F4")

put("G1", "RETAIL PURCHASE AGREEMENT", size=14, bold=True, align="center", merge="G1:L2")

put("K3", "CUST#", size=8, align="right")
put("L3", None, size=9, border=underline)
put("K4", "Deal Number:", size=8, align="right")
put("L4", None, size=9, border=underline)

for r, h in ((1, 20), (2, 14), (3, 13), (4, 13), (5, 6)):
    ws.row_dimensions[r].height = h

# ------------------------------------------------------- purchaser block
for r, left_lbl, right_lbl in ((6, "Purchaser's Name(s):", "Date:"),
                               (7, "Address(es):", "County:"),
                               (8, "Telephone (1):", "Telephone (2):"),
                               (9, "Email (1):", "Email (2):")):
    ws.row_dimensions[r].height = 16
    put("B%d" % r, left_lbl, size=8, merge="B%d:C%d" % (r, r))
    put("D%d" % r, None, size=9, merge="D%d:H%d" % (r, r), border=underline)
    put("K%d" % r, right_lbl, size=8, align="right")
    put("L%d" % r, None, size=9, border=underline)

# ------------------------------------------------------- manufacturer emblem frame
ws.merge_cells("I3:J9")
outline("I3:J9")
img = XLImage(EMBLEM)
img.width, img.height = 118, 44
marker = AnchorMarker(col=8, colOff=2 * EMU, row=5, rowOff=0)  # col I, row 6: centred in I3:J9
img.anchor = OneCellAnchor(_from=marker,
                           ext=XDRPositiveSize2D(img.width * EMU, img.height * EMU))
ws.add_image(img)

# ------------------------------------------------------- odometer notice
put("B10",
    "The Odometer Reading for the Vehicle You are purchasing is accurate unless indicated otherwise. "
    "Please refer to the Odometer Mileage Statement for full disclosure.",
    size=6.5, wrap=True, merge="B10:L10")
ws.row_dimensions[10].height = 20
ws.row_dimensions[11].height = 5

# ------------------------------------------------------- vehicle table
put("B12", "YEAR", size=7.5, bold=True, align="center")
put("C12", "MAKE", size=7.5, bold=True, align="center")
put("D12", "MODEL", size=7.5, bold=True, align="center", merge="D12:E12")
put("F12", "COLOR", size=7.5, bold=True, align="center", merge="F12:G12")
put("H12", "STOCK NO.", size=7.5, bold=True, align="center", merge="H12:I12")
ws.row_dimensions[12].height = 13

for rng in ("B13:B13", "C13:C13", "D13:E13", "F13:G13", "H13:I13"):
    a, b = rng.split(":")
    put(a, None, size=9, align="center", merge=None if a == b else rng)
    outline(rng)
ws.row_dimensions[13].height = 18

put("B14", "VIN", size=7.5, bold=True, align="center", merge="B14:C14")
put("D14", "(Vehicle)", size=7, italic=True, align="center")
put("E14", "ODOMETER READING\n☐ Not Accurate", size=6.5, bold=True, align="center",
    wrap=True, merge="E14:F14")
put("G14", "SALESPERSON", size=7.5, bold=True, align="center", merge="G14:I14")
ws.row_dimensions[14].height = 18

for rng in ("B15:C15", "E15:F15", "G15:I15"):
    put(rng.split(":")[0], None, size=9, align="center", merge=rng)
    outline(rng)
ws.row_dimensions[15].height = 18

put("B16", "THE VEHICLE IS:\n☐ NEW      ☐ USED", size=7, bold=True, wrap=True, merge="B16:D16")
put("E16", "PRIOR USE DISCLOSURE:   ☐ DEMONSTRATOR   ☐ PREVIOUSLY LEASED   "
           "☐ EXECUTIVE VEHICLE   ☐ RENTAL   ☐ OTHER ______",
    size=6, bold=True, wrap=True, merge="E16:I16")
outline("B16:I16")
ws.row_dimensions[16].height = 22

# ------------------------------------------------------- left column: statements
WARRANTY = ("We are selling this Vehicle to You AS-IS and We expressly disclaim all warranties, express and "
            "implied, including any implied warranties of merchantability and fitness for a particular purpose, "
            "unless the box beside \"Used Vehicle Limited Warranty Applies\" is marked below or We enter into a "
            "service contract with You at the time of, or within 90 days of, the date of this transaction. All "
            "warranties, if any, by a manufacturer or supplier other than Our Dealership are theirs, not Ours, and "
            "no manufacturer or supplier shall be liable for performance under such warranties. We neither assume "
            "nor authorize any other person to assume for Us any liability in connection with the sale of the "
            "Vehicle and related goods and services.")
DISCLOSURE = ("You see on the window form for this Vehicle is part of this contract. Information on the window form "
              "overrides any contrary provisions in the contract of sale.  Traducción española: Guía para "
              "compradores de vehículos usados. La información que ve en el formulario de la ventanilla para este "
              "vehículo forma parte del presente contrato. La información del formulario de la ventanilla deja sin "
              "efecto toda disposición en contrario contenida en el contrato de venta.")
UVLW = ("☐ Used Vehicle Limited Warranty Applies:  We are providing a Used Vehicle Limited Warranty in "
        "connection with this transaction. Any implied warranties apply for the duration of the Limited Warranty.")
INV_TAX = ("Dealer's Inventory Tax: The Dealer's Inventory Tax charge is intended to reimburse the Dealer for ad "
           "valorem taxes on its motor vehicle inventory. The charge, which is paid by the Dealer to the county tax "
           "assessor-collector, is not a tax imposed on a consumer by the government, and is not required to be "
           "charged to the consumer.")
DOC_FEE = ("*Documentary Fee: A documentary fee is not an official fee. A documentary fee is not required by law, "
           "but may be charged to buyers for handling documents relating to the sale. A documentary fee may not "
           "exceed a reasonable amount agreed to by the parties. This notice is required by law.  "
           "Traducción española: Vea Párrafo 13.")

put("B17", "WARRANTY STATEMENT", size=10, bold=True, align="center", merge="B17:I17")
outline("B17:I17")
ws.row_dimensions[17].height = 15

put("B18", WARRANTY, size=7, wrap=True, valign="top", merge="B18:I25")
outline("B18:I25")

put("B26", "CONTRACTUAL DISCLOSURE STATEMENT (USED VEHICLES ONLY) The information",
    size=7, bold=True, merge="B26:I26")
put("B27", DISCLOSURE, size=7, wrap=True, valign="top", merge="B27:I31")
outline("B26:I31")
ws.row_dimensions[26].height = 13

put("B32", UVLW, size=7, wrap=True, valign="top", merge="B32:I34")
outline("B32:I34")

for r in list(range(18, 26)) + list(range(27, 35)):
    ws.row_dimensions[r].height = 11

# --- trade vehicle information
put("B35", "TRADE VEHICLE INFORMATION", size=10, bold=True, align="center", merge="B35:I35")
outline("B35:I35")
ws.row_dimensions[35].height = 15


def trade_block(r0, label):
    put("B%d" % r0, "Lienholder Name:", size=7, merge="B%d:C%d" % (r0, r0))
    put("D%d" % r0, None, size=8, merge="D%d:I%d" % (r0, r0), border=underline)
    put("B%d" % (r0 + 1), "Lienholder Address:", size=7, merge="B%d:C%d" % (r0 + 1, r0 + 1))
    put("D%d" % (r0 + 1), None, size=8, merge="D%d:I%d" % (r0 + 1, r0 + 1), border=underline)

    r = r0 + 2
    for col, lbl in (("B", "Year:"), ("D", "Make:"), ("F", "Model:"), ("H", "Color:")):
        put("%s%d" % (col, r), lbl, size=7)
    for col in ("C", "E", "G", "I"):
        put("%s%d" % (col, r), None, size=8, border=underline)

    r = r0 + 3
    put("B%d" % r, "VIN:", size=7)
    put("C%d" % r, None, size=8, merge="C%d:D%d" % (r, r), border=underline)
    put("E%d" % r, "Odometer:", size=6.5)
    put("F%d" % r, None, size=8, border=underline)
    put("G%d" % r, "☐ Not Accurate        %s" % label, size=6.5, merge="G%d:I%d" % (r, r))

    r = r0 + 4
    put("B%d" % r, "Trade Allowance:", size=7, merge="B%d:C%d" % (r, r))
    put("D%d" % r, None, size=8, fmt=MONEY, border=underline)
    put("E%d" % r, "Balance Owed to Lienholder:", size=7, merge="E%d:G%d" % (r, r))
    put("H%d" % r, None, size=8, fmt=MONEY, merge="H%d:I%d" % (r, r), border=underline)

    for rr in range(r0, r0 + 5):
        ws.row_dimensions[rr].height = 12.5


trade_block(36, "(Trade Vehicle 1)")
trade_block(41, "(Trade Vehicle 2)")
outline("B36:I45")

# --- other material understandings
put("B46", "OTHER MATERIAL UNDERSTANDINGS AND INTEGRATED DOCUMENTS",
    size=9, bold=True, align="center", merge="B46:I46")
outline("B46:I46")
ws.row_dimensions[46].height = 15

put("B47", "☐  PLEASE SEE THE DELIVERY CONFIRMATION", size=7.5, merge="B47:I47")
put("B48", "☐  PLEASE SEE THE CONDITIONAL DELIVERY AGREEMENT", size=7.5, merge="B48:I48")
put("B49", "Prospective Lienholder:", size=7.5, merge="B49:C49")
put("D49", None, size=8, merge="D49:I49", border=underline)
outline("B47:I49")
for r in (47, 48, 49):
    ws.row_dimensions[r].height = 13

put("B50", INV_TAX, size=7, bold=True, wrap=True, valign="top", merge="B50:I53")
put("B54", DOC_FEE, size=7, bold=True, wrap=True, valign="top", merge="B54:I57")
for r in range(50, 58):
    ws.row_dimensions[r].height = 11

# ------------------------------------------------------- right column: pricing
R_BASE, R_OPT_HDR = 17, 18
R_OPT_FIRST, R_OPT_LAST = 19, 37
R_TOTAL_SELL, R_TRADE_ALLOW, R_SUBTOTAL = 38, 39, 40
R_FEE_FIRST, R_FEE_LAST = 41, 50
R_TOTAL_DUE, R_DOWN, R_REBATES, R_CASH_DUE, R_FINANCED = 52, 53, 54, 55, 57

GREY = PatternFill("solid", fgColor="EFEFEF")


def price_row(r, label, bold=False, formula=None, shade=False):
    put("K%d" % r, label, size=7.5, bold=bold, shrink=True)
    put("L%d" % r, formula, size=9, bold=bold, align="right", fmt=MONEY)
    for cell in ("K%d" % r, "L%d" % r):
        ws[cell].border = box
        if shade:
            ws[cell].fill = GREY


price_row(R_BASE, "Base Selling Price", bold=True)
put("K%d" % R_OPT_HDR, "Optional Items:", size=7.5, italic=True)
for cell in ("K%d" % R_OPT_HDR, "L%d" % R_OPT_HDR):
    ws[cell].border = box
for r in range(R_OPT_FIRST, R_OPT_LAST + 1):
    price_row(r, None)

price_row(R_TOTAL_SELL, "TOTAL SELLING PRICE", bold=True, shade=True,
          formula="=L%d+SUM(L%d:L%d)" % (R_BASE, R_OPT_FIRST, R_OPT_LAST))
price_row(R_TRADE_ALLOW, "LESS: TRADE IN ALLOWANCE")
price_row(R_SUBTOTAL, "SUBTOTAL", bold=True, shade=True,
          formula="=L%d-L%d" % (R_TOTAL_SELL, R_TRADE_ALLOW))

for i, name in enumerate(["SALES TAX", "DEALER'S INVENTORY TAX", "DOCUMENTARY FEE*",
                          "STATE INSPECTION FEE", "DEPUTY SERVICE FEE", "LICENSE FEE",
                          "TITLE FEE", "City Road & Bridge", "Processing Fee",
                          "Govt Vehicle Inspection Replacement Fee"]):
    price_row(R_FEE_FIRST + i, name)

price_row(R_TOTAL_DUE, "TOTAL DUE", bold=True, shade=True,
          formula="=L%d+SUM(L%d:L%d)" % (R_SUBTOTAL, R_FEE_FIRST, R_FEE_LAST))
price_row(R_DOWN, "DOWN PAYMENT")
price_row(R_REBATES, "REBATES")
price_row(R_CASH_DUE, "LESS CASH DUE AT DELIVERY")
price_row(R_FINANCED, "AMOUNT TO BE FINANCED", bold=True, shade=True,
          formula="=L%d-L%d-L%d-L%d" % (R_TOTAL_DUE, R_DOWN, R_REBATES, R_CASH_DUE))

# ------------------------------------------------------- approvals & consent
ws.row_dimensions[58].height = 7
for col in "BCDEFGHIJKL":
    ws["%s59" % col].border = Border(top=medium)

ws.row_dimensions[59].height = 24
put("B59", "Customer Approval:", size=8, bold=True, valign="bottom", merge="B59:C59")
put("D59", None, size=10, valign="bottom", merge="D59:F59")
put("G59", "Management Approval:", size=8, bold=True, valign="bottom", merge="G59:I59")
put("J59", None, size=10, valign="bottom", merge="J59:L59")
for col in ("D", "E", "F", "J", "K", "L"):
    b = ws["%s59" % col].border
    ws["%s59" % col].border = Border(top=b.top, bottom=thin)

CONSENT = ("By signing this authorization form, you certify that the above personal information is correct and "
           "accurate, and authorize the release of credit and employment information. By signing above, I provide "
           "to the dealership and its affiliates consent to communicate with me about my vehicle or any future "
           "vehicles using electronic verbal and written communications including but not limited to email, text "
           "messaging, SMS, phone calls and direct mail. Terms and conditions subject to credit approval. For "
           "information only. This is not an offer or contract for sale.")
put("B60", CONSENT, size=6.5, wrap=True, valign="top", merge="B60:L63")
for r in range(60, 64):
    ws.row_dimensions[r].height = 9
outline("B59:L63")

put("B64", "IMPORTANT TERMS AND CONDITIONS FOLLOW", size=9, bold=True, merge="B64:G64")
ws.row_dimensions[64].height = 16

# ------------------------------------------------------- page setup (A4, full page)
ws.print_area = "A1:M64"
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins = PageMargins(left=0.22, right=0.22, top=0.2, bottom=0.2,
                              header=0.1, footer=0.1)
ws.sheet_view.showGridLines = False

# ======================================================= instructions sheet
gs = wb.create_sheet("使い方 (Instructions)")
for col, w in (("A", 24), ("B", 66), ("C", 24)):
    gs.column_dimensions[col].width = w


def g(row, a, b, c=None, bold=False, size=10):
    gs.cell(row=row, column=1, value=a).font = Font(name=FONT, size=size, bold=bold)
    cb = gs.cell(row=row, column=2, value=b)
    cb.font = Font(name=FONT, size=size, bold=bold)
    cb.alignment = Alignment(wrap_text=True, vertical="top")
    if c is not None:
        gs.cell(row=row, column=3, value=c).font = Font(name=FONT, size=size, bold=bold)


g(1, "RETAIL PURCHASE AGREEMENT", "空白フォーム / Blank form - 記入ガイド (A4 1ページ)", bold=True, size=12)
g(3, "入力欄 (Input)", "説明", "記入例 (Example)", bold=True)
rows = [
    ("B1:F4", "ディーラー名・住所・Tel / Fax", "EXCELLENT AUTO SALES..."),
    ("I3:J9", "自動車メーカーのエンブレム枠 (画像を差し替え)", "Ford エンブレム"),
    ("L3 / L4", "CUST# / Deal Number", "408132 / 343179"),
    ("D6:H6 / L6", "Purchaser's Name(s) / Date", "JOHN DOE / 08/06/2026"),
    ("D7:H7 / L7", "Address(es) / County", "541 RUBY RIVER DR / MADISON"),
    ("D8:H8 / L8", "Telephone (1) / Telephone (2)", "555-0100 / N/A"),
    ("D9:H9 / L9", "Email (1) / Email (2)", "sample@example.com"),
    ("B13:I13", "YEAR / MAKE / MODEL / COLOR / STOCK NO.", "2026 / FORD / BRONCO"),
    ("B15:I15", "VIN / ODOMETER READING / SALESPERSON", "1FMEE4DP3TLB00730 / 11"),
    ("B16:I16", "New・Used、Prior Use Disclosure のチェック欄 (☐ を ☑ に置換)", "☑ NEW"),
    ("C36:I45", "TRADE VEHICLE INFORMATION (下取車 1・2)", "N/A"),
    ("D49", "Prospective Lienholder", "N/A"),
    ("L17", "Base Selling Price", "$52,810.00"),
    ("K19:L37", "Optional Items - 品名を K 列、金額を L 列に入力", "ALARM-AT / $305.00"),
    ("L39", "LESS: TRADE IN ALLOWANCE", "$0.00"),
    ("L41:L50", "各種税金・手数料 (SALES TAX 〜 Govt Vehicle Inspection Replacement Fee)", "$225.00"),
    ("L53 / L54 / L55", "DOWN PAYMENT / REBATES / LESS CASH DUE AT DELIVERY", "$0.00 / $2,000.00 / $0.00"),
    ("D59 / J59", "Customer Approval / Management Approval (署名欄)", "署名"),
]
r = 4
for a, b, c in rows:
    g(r, a, b, c)
    r += 1

r += 1
g(r, "自動計算 (編集しないこと)", "計算式", "", bold=True)
r += 1
for a, b in (("L38  TOTAL SELLING PRICE", "= L17 + SUM(L19:L37)"),
             ("L40  SUBTOTAL", "= L38 - L39"),
             ("L52  TOTAL DUE", "= L40 + SUM(L41:L50)"),
             ("L57  AMOUNT TO BE FINANCED", "= L52 - L53 - L54 - L55")):
    g(r, a, b, "")
    r += 1

r += 1
for note in [
    "・用紙は A4 縦・1ページに収まるよう設定済み (余白 上下 0.2\" / 左右 0.22\")。そのまま印刷してください。",
    "・金額欄の書式は $#,##0.00 (USD)。マイナスは ($#,##0.00)、空欄は「-」と表示されます。",
    "・グレー網掛けの行は数式です。上書きすると再計算されません。",
    "・エンブレムは I3:J9 の枠内に画像として配置しています。差し替えは画像を選択して削除し、"
    "[挿入]→[画像] で同じ枠に貼り直してください。",
]:
    g(r, "", note, "")
    r += 1

wb.save(OUT)
print("saved", OUT)
